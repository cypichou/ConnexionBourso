from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import date
import os

##__________________________________________________## ALGO DE RECUP DE L'EXCEL ##__________________________________________________##

dossier_telechargements = os.path.expanduser('~/Downloads')
fichiers_telechargements = os.listdir(dossier_telechargements)

switch = True
for fichier in fichiers_telechargements:
    if fichier.startswith('export-operations'):
        if switch:
            chemin_fichier1 = os.path.join(dossier_telechargements, fichier)
            switch = not switch
        else:
            chemin_fichier2 = os.path.join(dossier_telechargements, fichier)

infos_fichier = os.stat(chemin_fichier1)
date_creation_fichier1 = infos_fichier.st_ctime

infos_fichier = os.stat(chemin_fichier2)
date_creation_fichier2 = infos_fichier.st_ctime

if date_creation_fichier1 > date_creation_fichier2:
    chemin_fichier_incomes = chemin_fichier1
    chemin_fichier_expenses = chemin_fichier2
else:
    chemin_fichier_incomes = chemin_fichier2
    chemin_fichier_expenses = chemin_fichier1

##__________________________________________________## CSV ##__________________________________________________##

def CsvToString(str) :
    dic={}
    tab=[]
    indice_dic=0
    retourLigne=0
    i=1
    actuel=''
    is_dic=True

    while i<len(str):
        
        while str[i]!=';' and str[i] != '\n':
            actuel+=str[i]
            i+=1
        
        if is_dic:
            dic[actuel] = indice_dic
            is_dic = False if str[i] == '\n' else True 
            if not is_dic : 
                tab.append([]) 
            indice_dic+=1      

        else:
            tab[retourLigne].append(actuel)
            if str[i]=='\n' and i<=len(str)-1:
                tab.append([])
                retourLigne+=1
            
        actuel=''
        i+=1
    
    return dic,tab[:len(tab)-1] 

colonne_dic = {}
contenu_tab = []

with open(chemin_fichier_incomes, encoding="utf-8") as csv:
    csv_str=csv.read()

colonne_dic,contenu_tab = CsvToString(csv_str)

##__________________________________________________## EXCEL ##__________________________________________________##

def delQuote(str,long):
       
   if str[0] == '"':
        return str[1:long-1].replace(' ','').replace(',','.')
   elif str[0] == '-':
        return str[1:].replace(' ','').replace(',','.').replace('-','')
   else: 
       return str.replace(' ','').replace(',','.')
   
def trouverLaCategorie(Supplier,category,categoryParent): 
    
    categorie = "nop"
    
    ws = wb['Bibliotheque']
    table = ws.tables['bibli']  
    derniere_ligne = int(table.ref[table.ref.find('C')+1:])

    trouve = False
    i=2

    while i in range(2,derniere_ligne+1) and trouve == False:
        if ws[f'A{i}'].value == Supplier:
            categorie = ws[f'B{i}'].value
            trouve = False
        i+=1    

    if categorie=="nop":
        ws.insert_rows(2)
        categorie = Supplier
        ws['A2'].value=Supplier
        ws['B2'].value=category
        ws['C2'].value=categoryParent

    return categorie

wb = load_workbook('C:\\Users\\cypri\\Documents\\Perso_Finances.xlsx')
ws = wb['Budget Tracking']

table = ws.tables['Tracking']  
derniere_ligne = int(table.ref[len(table.ref)-2:])+1

for j in range(2):

    for i in range(0,len(contenu_tab)):
        
        decoupage = contenu_tab[i][colonne_dic['dateOp']].split('/')
        ws[f"C{derniere_ligne}"].value = datetime(int(decoupage[2]),int(decoupage[1]),int(decoupage[0]))
        amount = float(delQuote(contenu_tab[i][colonne_dic['amount']], len(contenu_tab[i][colonne_dic['amount']])))

        if j != 0:

            ws[f"D{derniere_ligne}"].value = "Income"
            ws[f"E{derniere_ligne}"].value = trouverLaCategorie(contenu_tab[i][colonne_dic['supplierFound']],contenu_tab[i][colonne_dic['category']],contenu_tab[i][colonne_dic['categoryParent']])
            ws[f"F{derniere_ligne}"].value = amount
            wb.save('C:\\Users\\cypri\\Documents\\Perso_Finances.xlsx')

        else:

            ws[f"D{derniere_ligne}"].value = "Expenses"
            valueTest = trouverLaCategorie(contenu_tab[i][colonne_dic['supplierFound']],contenu_tab[i][colonne_dic['category']],contenu_tab[i][colonne_dic['categoryParent']])
            ws[f"E{derniere_ligne}"].value = valueTest
            ws[f"F{derniere_ligne}"].value = amount
        
        derniere_ligne+=1
            
    if j==0:
        colonne_dic = {}
        contenu_tab = []

        with open(chemin_fichier_expenses, encoding="utf-8") as csv:
            csv_str=csv.read()

        colonne_dic,contenu_tab = CsvToString(csv_str)



wb.save('C:\\Users\\cypri\\Documents\\Perso_Finances.xlsx')
# os.remove(chemin_fichier1)
# os.remove(chemin_fichier2)






