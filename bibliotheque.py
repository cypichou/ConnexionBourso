from openpyxl import Workbook, load_workbook


wb = load_workbook('C:\\Users\\cypri\\Documents\\Perso_Finances.xlsx')

ws = wb['Bibliotheque']
ws['D1'].value = "bite"
print(ws['A1'].value)



wb.save('C:\\Users\\cypri\\Documents\\Perso_Finances.xlsx')

